import pytest
import os, sys, inspect
import sys, os
myPath = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, myPath + '/../../')
from models.discreteWeibull2 import DiscreteWeibull2
from models.geometric import Geometric
from models.negativeBinomial2 import NegativeBinomial2
from core.dataClass import Data
import configparser
import openpyxl
import pandas as pd
import logging

logging.basicConfig(level=logging.DEBUG)
mylogger = logging.getLogger()

mylogger.info('\n###############\nStarting  Model Testing\n###############')

def DWSetup(Systemdata):
    MetricCombos = [[],['E'],['F'],['C'],['E','F'],['F','C'],['E','C'],['E','F','C']]
    DWresults = {}
    config = configparser.ConfigParser()
    config.read('config.ini')
    for metricname in MetricCombos:
        DW = DiscreteWeibull2(data=Systemdata.getFullData(),metricNames=metricname,config=config)
        DW.runEstimation(DW.covariateData)
        if (metricname == []):
            metricname = ['-']
        DWresults[listToString(metricname)] = {"b":DW.modelParameters[0],"Beta":DW.betas ,
                                               "Omega": DW.omega,"LL" : DW.llfVal}
    return DWresults

def GMSetup(Systemdata):
    MetricCombos = [[], ['E'], ['F'], ['C'], ['E', 'F'], ['F', 'C'], ['E', 'C'], ['E', 'F', 'C']]
    GMresults = {}
    config = configparser.ConfigParser()
    config.read('config.ini')
    for metricname in MetricCombos:
        GM = Geometric(data=Systemdata.getFullData(), metricNames=metricname, config=config)
        GM.runEstimation(GM.covariateData)
        if (metricname == []):
            metricname = ['-']
        GMresults[listToString(metricname)] = {"b": GM.modelParameters[0], "Beta": GM.betas,
                                               "Omega": GM.omega,"LL": GM.llfVal}
    return GMresults

def NBSetup(Systemdata):
    MetricCombos = [[], ['E'], ['F'], ['C'], ['E', 'F'], ['F', 'C'], ['E', 'C'], ['E', 'F', 'C']]
    NBresults = {}
    config = configparser.ConfigParser()
    config.read('config.ini')
    for metricname in MetricCombos:
        NB = NegativeBinomial2(data=Systemdata.getFullData(), metricNames=metricname, config=config)
        NB.runEstimation(NB.covariateData)
        if (metricname == []):
            metricname = ['-']
        NBresults[listToString(metricname)] = {"b": NB.modelParameters[0], "Beta": NB.betas,
                                               "Omega": NB.omega,"LL": NB.llfVal}
    return NBresults


def listToString(s):
    # initialize an empty string
    str1 = ""
    
    # traverse in the string
    for ele in s:
        str1 += ele
        
        # return string
    return str1

def ExtractExpectedMLE(sheetNum):
    fname = 'MLEs.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb[wb.sheetnames[sheetNum]]
    dic = {'DS1':{'-':{},'E':{},'F':{},'C':{},'EF':{},'FC':{},'EC':{},'EFC':{}},
           'DS2':{'-':{},'E':{},'F':{},'C':{},'EF':{},'FC':{},'EC':{},'EFC':{}}}


    ######## DS1 ###########
    for i in range(3,11):
        dic['DS1'][sheet.cell(i,1).value]['Omega'] =  sheet.cell(i,2).value
        dic['DS1'][sheet.cell(i,1).value]['b'] = sheet.cell(i,3).value
        dic['DS1'][sheet.cell(i,1).value]['LL'] = sheet.cell(i,7).value
        if i > 3 and i < 7 :
            dic['DS1'][sheet.cell(i, 1).value]['Beta'] = [sheet.cell(i, 4).value]
        elif i > 7 and i < 10 :
            dic['DS1'][sheet.cell(i, 1).value]['Beta'] = [sheet.cell(i, 4).value,sheet.cell(i, 5).value]
        else:
            dic['DS1'][sheet.cell(i, 1).value]['Beta'] = [sheet.cell(i, 4).value,sheet.cell(i, 5).value,sheet.cell(i, 6).value]
    for i in range(15,23):
        dic['DS2'][sheet.cell(i,1).value]['Omega'] =  sheet.cell(i,2).value
        dic['DS2'][sheet.cell(i,1).value]['b'] = sheet.cell(i,3).value
        dic['DS2'][sheet.cell(i,1).value]['LL'] = sheet.cell(i,7).value
        if i > 3 and i < 7 :
            dic['DS2'][sheet.cell(i, 1).value]['Beta'] = [sheet.cell(i, 4).value]
        elif i > 7 and i < 10 :
            dic['DS2'][sheet.cell(i, 1).value]['Beta'] = [sheet.cell(i, 4).value,sheet.cell(i, 5).value]
        else:
            dic['DS2'][sheet.cell(i, 1).value]['Beta'] = [sheet.cell(i, 4).value,sheet.cell(i, 5).value,sheet.cell(i, 6).value]
    return  dic


fname = "ds1.csv"
SystemdataDS1 = Data()
SystemdataDS1.importFile(fname)
DWresultsDS1 = DWSetup(SystemdataDS1)
GMresultsDS1 = GMSetup(SystemdataDS1)
NBresultsDS1 = NBSetup(SystemdataDS1)


fname = "ds2.csv"
SystemdataDS2 = Data()
SystemdataDS2.importFile(fname)
DWresultsDS2 = DWSetup(SystemdataDS2)
GMresultsDS2 = GMSetup(SystemdataDS2)
NBresultsDS2 = NBSetup(SystemdataDS2)

DWExpected = ExtractExpectedMLE(0)
GMExpected = ExtractExpectedMLE(1)
NBExpected = ExtractExpectedMLE(2)

DWtestomega = []
DWtestb = []
DWtestLL = []
DWtestBetas = []

for key in DWExpected['DS1']:
    DWtestomega.append((DWresultsDS1[key]['Omega'],DWExpected['DS1'][key]['Omega'],'DS1'))
    DWtestomega.append((DWresultsDS2[key]['Omega'], DWExpected['DS2'][key]['Omega'], 'DS2'))
for key in DWExpected['DS1']:
    DWtestb.append((DWresultsDS1[key]['b'],DWExpected['DS1'][key]['b'],'DS1'))
    DWtestb.append((DWresultsDS2[key]['b'], DWExpected['DS2'][key]['b'], 'DS2'))

for key in DWExpected['DS1']:
    DWtestLL.append((DWresultsDS1[key]['LL'], DWExpected['DS1'][key]['LL'], 'DS1'))
    DWtestLL.append((DWresultsDS2[key]['LL'], DWExpected['DS2'][key]['LL'], 'DS2'))
    

for key in DWExpected['DS1']:
    if key == '-':
        continue
    else:
        for i in range(0,len(DWresultsDS1[key]['Beta'])):
            DWtestBetas.append((DWresultsDS1[key]['Beta'][i],DWExpected['DS1'][key]['Beta'][i], 'DS1'))
            DWtestBetas.append((DWresultsDS2[key]['Beta'][i], DWExpected['DS2'][key]['Beta'][i], 'DS2'))


GMtestomega = []
GMtestb = []
GMtestLL = []
GMtestBetas = []

for key in GMExpected['DS1']:
    GMtestomega.append((GMresultsDS1[key]['Omega'],GMExpected['DS1'][key]['Omega'],'DS1'))
    GMtestomega.append((GMresultsDS2[key]['Omega'],GMExpected['DS2'][key]['Omega'], 'DS2'))
for key in GMExpected['DS1']:
    GMtestb.append((GMresultsDS1[key]['b'],GMExpected['DS1'][key]['b'],'DS1'))
    GMtestb.append((GMresultsDS2[key]['b'], GMExpected['DS2'][key]['b'], 'DS2'))

for key in GMExpected['DS1']:
    GMtestLL.append((GMresultsDS1[key]['LL'], GMExpected['DS1'][key]['LL'], 'DS1'))
    GMtestLL.append((GMresultsDS2[key]['LL'], GMExpected['DS2'][key]['LL'], 'DS2'))
    
    
for key in GMExpected['DS1']:
    if key == '-':
        continue
    else:
        for i in range(0,len(GMresultsDS1[key]['Beta'])):
            GMtestBetas.append((GMresultsDS1[key]['Beta'][i],GMExpected['DS1'][key]['Beta'][i], 'DS1'))
            GMtestBetas.append((GMresultsDS2[key]['Beta'][i], GMExpected['DS2'][key]['Beta'][i], 'DS2'))


NBtestomega = []
NBtestb = []
NBtestLL = []
NBtestBetas = []

for key in NBExpected['DS1']:
    NBtestomega.append((NBresultsDS1[key]['Omega'],NBExpected['DS1'][key]['Omega'],'DS1'))
    NBtestomega.append((NBresultsDS2[key]['Omega'],NBExpected['DS2'][key]['Omega'], 'DS2'))
for key in NBExpected['DS1']:
    NBtestb.append((NBresultsDS1[key]['b'],NBExpected['DS1'][key]['b'],'DS1'))
    NBtestb.append((NBresultsDS2[key]['b'], NBExpected['DS2'][key]['b'], 'DS2'))

for key in NBExpected['DS1']:
    NBtestLL.append((NBresultsDS1[key]['LL'], NBExpected['DS1'][key]['LL'], 'DS1'))
    NBtestLL.append((NBresultsDS2[key]['LL'], NBExpected['DS2'][key]['LL'], 'DS2'))

for key in NBExpected['DS1']:
    if key == '-':
        continue
    else:
        for i in range(0,len(NBresultsDS1[key]['Beta'])):
            NBtestBetas.append((NBresultsDS1[key]['Beta'][i],NBExpected['DS1'][key]['Beta'][i], 'DS1'))
            NBtestBetas.append((NBresultsDS2[key]['Beta'][i], NBExpected['DS2'][key]['Beta'][i], 'DS2'))



################Discrete Weibull###################
@pytest.mark.parametrize("expected , result,sheet",DWtestomega)
def test_DW_Omega_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet",DWtestb)
def test_DW_b_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet",DWtestLL)
def test_DW_LL_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet", DWtestBetas)
def test_GM_LL_mle(expected, result, sheet):
    assert abs(expected - result) < 10 ** -8


############Geometric###############

@pytest.mark.parametrize("expected , result,sheet",GMtestomega)
def test_GM_Omega_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet",GMtestb)
def test_GM_b_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet",GMtestLL)
def test_GM_LL_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet", GMtestBetas)
def test_GM_LL_mle(expected, result, sheet):
    assert abs(expected - result) < 10 ** -8
    





############Negative Binomial###############

@pytest.mark.parametrize("expected , result,sheet",NBtestomega)
def test_NB_Omega_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet",NBtestb)
def test_NB_b_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8

@pytest.mark.parametrize("expected , result,sheet",NBtestLL)
def test_NB_LL_mle(expected , result,sheet):
    assert abs(expected-result) < 10**-8
    
@pytest.mark.parametrize("expected , result,sheet", NBtestBetas)
def test_GM_LL_mle(expected, result, sheet):
    assert abs(expected - result) < 10 ** -8
